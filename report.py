import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from config import OUTPUT_CSV, OUTPUT_XLSX

def main():
    # 读取 CSV
    df = pd.read_csv(OUTPUT_CSV)

    # 保存为 Excel
    df.to_excel(OUTPUT_XLSX, index=False)

    wb = load_workbook(OUTPUT_XLSX)
    ws = wb.active

    # 冻结首行
    ws.freeze_panes = "A2"

    # 表头样式
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # 自动列宽
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

    wb.save(OUTPUT_XLSX)

    print("✅ Pretty Excel generated:", OUTPUT_XLSX)

if __name__ == "__main__":
    main()